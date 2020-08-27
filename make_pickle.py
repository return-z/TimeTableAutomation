from openpyxl import load_workbook
import pickle,re

class component:
    pass

wb = load_workbook('old.xlsx')
ws = wb['Table 1']
courseCodes = []
courseSpan = []
subjectWise = {}

for row in ws.iter_rows():
    try:
        if int(row[0].value):
            courseSpan.append(int(row[0].row))
    except:
        pass

for i in range(len(courseSpan)-1):
    code = f'B{courseSpan[i]}'
    title = f'C{courseSpan[i]}'
    subjectWise[(f"{ws[code].value}", f"{ws[title].value}")] = key = []
    start = courseSpan[i]
    end = courseSpan[i+1] - 1
    for row in ws.iter_rows(min_row=start, max_row=end):
        if row[8].value and row[8].value != "ROOM":
            key.append(row[8].row)
    key.append(courseSpan[i+1])

objects = {}

def correctedHour(ws, cell):
    hour = str(ws[cell].value)
    try:
        if len(hour) == 2 and int(hour) > 11:
            return [int(hour[0]), int(hour[1])]
        if len(hour) == 3:
            return [int(hour[0]), int(hour[1:])]
        if len(hour) == 4:
            return [int(hour[:2]), int(hour[2:])]
        else:
            return [int(hour)]
    except:
        return [ws[cell].value]

for subject in subjectWise:
    objects[subject] = {'Lecture' : [], 'Tutorial' : [], 'Practical' : [], 'Misc' : []}
    arr = subjectWise[subject]
    cell = f'I{arr[0]}' # for info regarding room number

    if len(arr) == 1:
        courseObject = component()
        courseObject.room = ws[cell].value
        courseObject.days = ws[f'J{arr[0]}'].value
        courseObject.hours =  correctedHour(ws, f'K{arr[0]}')
        courseObject.section = ws[f'G{arr[0]}'].value
        for row in ws.iter_rows(min_row=arr[0], max_row=arr[0]):
            courseObject.teachers = (row[7].value)
        objects[subject]['Misc'].append(courseObject)
    courseType = "Lecture"

    for i in range(len(arr)-1):
        courseObject = component()
        courseObject.teachers = []
        courseObject.room = ws[cell].value
        courseObject.days = re.split(r'\s', ws[f'J{arr[i]}'].value)
        courseObject.hours =  correctedHour(ws, f'K{arr[i]}')
        courseObject.section = ws[f'G{arr[i]}'].value
        if ws[f'C{arr[i]}'].value == "Tutorial" or ws[f'C{arr[i]}'].value == "Practical":
            courseType = ws[f'C{arr[i]}'].value
        for row in ws.iter_rows(min_row=arr[i], max_row=arr[i+1]-1):
            if not row[7].value == "INSTRUCTOR-IN-CHARGE/\nInstructor":
                courseObject.teachers.append(row[7].value)
        objects[subject][courseType].append(courseObject)

def save_obj():
    with open("objects" + '.pkl', 'wb') as f:
        pickle.dump(objects, f, pickle.HIGHEST_PROTOCOL)

save_obj()


