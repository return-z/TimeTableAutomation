from openpyxl import load_workbook
import pickle

streams = {}
wb = load_workbook('cdcs.xlsx')

sheets = wb.sheetnames

for sheet in sheets:
    ws = wb[sheet]
    span = []
    for row in ws.iter_rows():
        if row[0].value:
            span.append(row[0].row)
    span.append(ws.max_row+1)
    streams[sheet] = span

cdc = {}

for stream in streams:
    ws = wb[stream]
    cdc[stream] = {}
    arr = streams[stream]
    for i in range(len(arr)-1):
        year = ws[f'A{arr[i]}'].value
        sem = ws[f'B{arr[i]}'].value
        cdc[stream][(year, sem)] = subjects = []
        for row in ws.iter_rows(min_row=arr[i], max_row=arr[i+1]-1):
            subjects.append(row[2].value)

def save_obj():
    with open("cdc" + '.pkl', 'wb') as f:
        pickle.dump(cdc, f,protocol=3)

save_obj()
